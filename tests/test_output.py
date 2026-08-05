import os
import tempfile
import unittest

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from src.anomaly_detection.output import build_output, write_output


class BuildOutputTests(unittest.TestCase):
    def test_build_output_sanitizes_columns_and_adds_interpretable_fields(self):
        df = pd.DataFrame(
            {
                "facility id": ["A", "B"],
                "metric/one": [1.0, 2.0],
                "metric two": [3.0, 4.0],
            }
        )

        result = build_output(
            df_processed=df,
            id_cols=["facility id"],
            distances_sq=np.array([1.0, 4.0]),
            estimates=np.array([[1.0, 1.0], [2.0, 2.0]]),
            deviations=np.array([[0.1, 0.2], [0.3, 0.4]]),
            outlier_flags=np.array([0, 1]),
            min_value_threshold=0.0,
        )

        self.assertIn("facility_id", result.columns)
        self.assertNotIn("facility id", result.columns)
        self.assertNotIn("metric/one", result.columns)
        self.assertNotIn("metric two", result.columns)
        self.assertNotIn("D_metric_one", result.columns)
        self.assertNotIn("D_metric_two", result.columns)

        self.assertIn("mahalanobis_distance", result.columns)
        self.assertIn("anomalous", result.columns)
        self.assertTrue(np.allclose(np.sort(result["mahalanobis_distance"].to_numpy()), np.sqrt([1.0, 4.0])))
        self.assertTrue(result.iloc[0]["anomalous"] or result.iloc[1]["anomalous"])

    def test_write_output_writes_excel_workbook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = os.path.join(tmpdir, "results.xlsx")
            df = pd.DataFrame({"facility": ["A"], "var1": [1.5], "var2": [2.5]})
            result = build_output(
                df_processed=df,
                id_cols=["facility"],
                distances_sq=np.array([1.0]),
                estimates=np.array([[1.0, 2.0]]),
                deviations=np.array([[0.1, 0.2]]),
                outlier_flags=np.array([1]),
                min_value_threshold=0.0,
            )
            report = {
                "metric_cols_used": ["var1", "var2"],
                "analysis_data": np.array([[1.0, 2.0], [3.0, 4.0]]),
                "covariance_matrix": np.array([[4.0, 2.0], [2.0, 9.0]]),
                "deviations": np.array([[0.5, 1.5]]),
            }
            written = write_output(result, report, out_path)
            self.assertEqual(written, out_path)
            self.assertTrue(os.path.exists(out_path))

            workbook = pd.read_excel(out_path, sheet_name="facility_results")
            self.assertIn("facility", workbook.columns)
            self.assertIn("var1", workbook.columns)
            self.assertIn("var2", workbook.columns)
            self.assertEqual(workbook.iloc[0]["var1"], "1.5 (exp 1)")
            self.assertEqual(workbook.iloc[0]["var2"], "2.5 (exp 2)")

            xlsx_workbook = load_workbook(out_path)
            results_sheet = xlsx_workbook["facility_results"]
            self.assertEqual(results_sheet["B2"].fill.fgColor.rgb, "00FF0000")
            self.assertEqual(results_sheet["C2"].fill.fgColor.rgb, "00000000")

            correlation_sheet = pd.read_excel(out_path, sheet_name="correlation_matrix", index_col=0)
            self.assertAlmostEqual(correlation_sheet.loc["var1", "var2"], 1.0 / 3.0, places=6)

    def test_small_expected_differences_are_not_colored(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = os.path.join(tmpdir, "results.xlsx")
            result = pd.DataFrame({
                "facility": ["A"],
                "var1": ["0 (exp 0.2)"],
                "var2": ["5 (exp 1.0)"],
                "mahalanobis_distance": [2.0],
                "anomalous": [True],
                "top_deviation_indicator": ["var2"],
            })
            report = {
                "metric_cols_used": ["var1", "var2"],
                "analysis_data": np.array([[0.0, 5.0]]),
                "covariance_matrix": np.array([[1.0, 0.0], [0.0, 1.0]]),
                "deviations": np.array([[0.2, 3.0]]),
                "id_cols": ["facility"],
            }
            write_output(result, report, out_path)

            xlsx_workbook = load_workbook(out_path)
            results_sheet = xlsx_workbook["facility_results"]
            self.assertEqual(results_sheet["B2"].fill.fgColor.rgb, "00FFFFFF")
            self.assertEqual(results_sheet["C2"].fill.fgColor.rgb, "00FF0000")

    def test_coloring_uses_source_row_alignment_after_sort(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = os.path.join(tmpdir, "results.xlsx")
            df = pd.DataFrame({"facility": ["A", "B"], "var1": [1.0, 10.0], "var2": [1.0, 10.0]})
            result = build_output(
                df_processed=df,
                id_cols=["facility"],
                distances_sq=np.array([1.0, 9.0]),
                estimates=np.array([[1.0, 1.0], [10.0, 2.0]]),
                deviations=np.array([[2.0, 0.1], [0.2, 3.0]]),
                outlier_flags=np.array([0, 1]),
                min_value_threshold=0.0,
            )
            report = {
                "metric_cols_used": ["var1", "var2"],
                "analysis_data": np.array([[1.0, 1.0], [10.0, 10.0]]),
                "estimated_data": np.array([[1.0, 1.0], [10.0, 2.0]]),
                "covariance_matrix": np.array([[1.0, 0.0], [0.0, 1.0]]),
                "deviations": np.array([[2.0, 0.1], [0.2, 3.0]]),
                "id_cols": ["facility"],
            }
            write_output(result, report, out_path)

            xlsx_workbook = load_workbook(out_path)
            results_sheet = xlsx_workbook["facility_results"]

            # Row 2 is the anomalous "B" row after sorting; var2 should be highlighted.
            self.assertEqual(results_sheet["A2"].value, "B")
            self.assertEqual(results_sheet["B2"].fill.fgColor.rgb, "00FFFFFF")
            self.assertEqual(results_sheet["C2"].fill.fgColor.rgb, "00FF0000")

    def test_zero_actual_large_expected_is_colored(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = os.path.join(tmpdir, "results.xlsx")
            result = pd.DataFrame({
                "facility": ["A"],
                "var1": ["0 (exp 500)"],
                "var2": ["10 (exp 10)"],
                "mahalanobis_distance": [5.0],
                "anomalous": [True],
                "top_deviation_indicator": ["var1"],
            })
            report = {
                "metric_cols_used": ["var1", "var2"],
                "analysis_data": np.array([[0.0, 10.0]]),
                "estimated_data": np.array([[500.0, 10.0]]),
                "covariance_matrix": np.array([[1.0, 0.0], [0.0, 1.0]]),
                "deviations": np.array([[2.0, 0.1]]),
                "id_cols": ["facility"],
            }
            write_output(result, report, out_path)

            xlsx_workbook = load_workbook(out_path)
            results_sheet = xlsx_workbook["facility_results"]
            self.assertEqual(results_sheet["B2"].fill.fgColor.rgb, "00FF0000")

    def test_anomalous_row_has_between_one_and_three_highlights(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = os.path.join(tmpdir, "results.xlsx")
            result = pd.DataFrame({
                "facility": ["A"],
                "var1": ["100 (exp 10)"],
                "var2": ["80 (exp 8)"],
                "var3": ["60 (exp 6)"],
                "var4": ["40 (exp 4)"],
                "mahalanobis_distance": [9.0],
                "anomalous": [True],
                "top_deviation_indicator": ["var1"],
            })
            report = {
                "metric_cols_used": ["var1", "var2", "var3", "var4"],
                "analysis_data": np.array([[100.0, 80.0, 60.0, 40.0]]),
                "estimated_data": np.array([[10.0, 8.0, 6.0, 4.0]]),
                "covariance_matrix": np.eye(4),
                "deviations": np.array([[4.0, 3.0, 2.0, 1.0]]),
                "id_cols": ["facility"],
            }
            write_output(result, report, out_path)

            xlsx_workbook = load_workbook(out_path)
            results_sheet = xlsx_workbook["facility_results"]
            highlighted = 0
            for col in ["B", "C", "D", "E"]:
                fill = results_sheet[f"{col}2"].fill.fgColor.rgb
                if fill not in ("00FFFFFF", "00000000", None):
                    highlighted += 1
            self.assertGreaterEqual(highlighted, 1)
            self.assertLessEqual(highlighted, 3)


if __name__ == "__main__":
    unittest.main()
