import os
import re

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
from scipy import linalg

from .covariance import stable_inv


def _sanitize_column_name(name: object) -> str:
    """Convert column names to Excel-friendly labels with only letters, numbers, and underscores."""
    cleaned = re.sub(r"[^0-9A-Za-z]+", "_", str(name)).strip("_").lower()
    return cleaned or "column"


def _sanitize_column_names(names: list) -> list:
    """Make a list of column names unique and Excel-friendly."""
    seen = {}
    sanitized = []
    for name in names:
        base = _sanitize_column_name(name)
        count = seen.get(base, 0)
        seen[base] = count + 1
        if count:
            sanitized.append(f"{base}_{count}")
        else:
            sanitized.append(base)
    return sanitized


def estimate_values(data: np.ndarray, mu: np.ndarray, cov: np.ndarray) -> np.ndarray:
    """
    For each present value in each row, estimate the conditional expectation
    given all other present values in that row:
      E[x_j | x_{-j}] = mu_j + cov[j, -j] @ inv(cov[-j, -j]) @ (x_{-j} - mu[-j])

    Returns array of same shape as data; NaN where the original value was missing.
    """
    n, k = data.shape
    estimates = np.full((n, k), np.nan)

    for i in range(n):
        inds = np.where(~np.isnan(data[i]))[0]
        if len(inds) < 2:
            continue
        for j in inds:
            other = inds[inds != j]
            cov_j_other = cov[j, other]
            cov_other = cov[np.ix_(other, other)]
            try:
                inv_cov_other = stable_inv(cov_other)
                diff = data[i, other] - mu[other]
                estimates[i, j] = mu[j] + cov_j_other @ inv_cov_other @ diff
            except linalg.LinAlgError:
                estimates[i, j] = mu[j]

    return estimates


def compute_deviations(data: np.ndarray, estimates: np.ndarray, cov: np.ndarray) -> np.ndarray:
    """
    Normalized absolute deviation per cell: |actual - estimated| / var[j].
    NaN where the original value was missing.
    """
    variances = np.diag(cov)
    with np.errstate(divide="ignore", invalid="ignore"):
        deviations = np.abs(data - estimates) / variances
    return deviations


def _fill_for_value(value: float, max_abs: float) -> PatternFill:
    if not np.isfinite(value) or max_abs == 0:
        return PatternFill(fill_type="solid", fgColor="FFFFFF")
    ratio = abs(float(value)) / max_abs
    if float(value) >= 0:
        intensity = int(255 * min(ratio, 1.0))
        return PatternFill(fill_type="solid", fgColor=f"FF{255-intensity:02x}{255-intensity:02x}")
    intensity = int(255 * min(ratio, 1.0))
    return PatternFill(fill_type="solid", fgColor=f"{255-intensity:02x}{255-intensity:02x}FF")


def _parse_numeric_value(value: object) -> float:
    if value is None:
        return np.nan
    if isinstance(value, (int, float)) and np.isfinite(value):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return np.nan
        if " (exp " in text:
            text = text.split(" (exp ", 1)[0]
        if text.startswith("(") and text.endswith(")"):
            text = text[1:-1].strip()
        try:
            return float(text)
        except ValueError:
            return np.nan
    return np.nan


def _parse_expected_from_display(value: object) -> float:
    if not isinstance(value, str):
        return np.nan
    text = value.strip()
    if " (exp " not in text:
        return np.nan
    expected_part = text.split(" (exp ", 1)[1].rstrip(")").strip()
    try:
        return float(expected_part)
    except ValueError:
        return np.nan


def _blended_driver_scores(
    row_devs: np.ndarray,
    row_actuals: np.ndarray,
    row_expecteds: np.ndarray,
    minimum_signal_scale: float,
    minimum_absolute_delta: float,
    relative_delta_floor: float,
) -> tuple[np.ndarray, np.ndarray]:
    """Blend normalized deviation and absolute residual into one driver score."""
    row_devs = np.asarray(row_devs, dtype=float)
    row_actuals = np.asarray(row_actuals, dtype=float)
    row_expecteds = np.asarray(row_expecteds, dtype=float)

    finite_mask = np.isfinite(row_devs) & np.isfinite(row_actuals) & np.isfinite(row_expecteds)
    abs_delta = np.abs(row_actuals - row_expecteds)
    abs_delta[~finite_mask] = np.nan

    max_dev = np.nanmax(row_devs) if np.any(np.isfinite(row_devs)) else 0.0
    max_delta = np.nanmax(abs_delta) if np.any(np.isfinite(abs_delta)) else 0.0

    dev_component = np.zeros_like(row_devs, dtype=float)
    delta_component = np.zeros_like(row_devs, dtype=float)
    if max_dev > 0:
        dev_component = np.divide(row_devs, max_dev, out=np.zeros_like(row_devs, dtype=float), where=np.isfinite(row_devs))
    if max_delta > 0:
        delta_component = np.divide(abs_delta, max_delta, out=np.zeros_like(row_devs, dtype=float), where=np.isfinite(abs_delta))

    # Keep normalized deviation primary while still rewarding large raw-unit residuals.
    score = 0.65 * dev_component + 0.35 * delta_component

    relative_floor = max_delta * max(relative_delta_floor, 0.0)
    delta_floor = max(minimum_absolute_delta, relative_floor)
    signal_scale = np.maximum(np.abs(row_actuals), np.abs(row_expecteds))
    eligible_mask = (
        finite_mask
        & (signal_scale >= minimum_signal_scale)
        & (abs_delta >= delta_floor)
    )
    score = np.where(eligible_mask, score, 0.0)
    return score, abs_delta


def _select_driver_indices(
    blended_scores: np.ndarray,
    finite_mask: np.ndarray,
    primary_mask: np.ndarray,
    is_anomalous: bool,
    min_count: int,
    max_count: int,
) -> list[int]:
    """Pick a compact set of driver indices for coloring."""
    if max_count <= 0:
        return []

    score_order = np.argsort(np.nan_to_num(blended_scores, nan=-np.inf))[::-1]
    selected = [int(idx) for idx in score_order if primary_mask[idx]]

    if is_anomalous and len(selected) < min_count:
        fallback = [
            int(idx)
            for idx in score_order
            if finite_mask[idx] and blended_scores[idx] > 0 and int(idx) not in selected
        ]
        selected.extend(fallback)

    if is_anomalous and len(selected) < min_count:
        last_resort = [int(idx) for idx in score_order if finite_mask[idx] and int(idx) not in selected]
        selected.extend(last_resort)

    if not is_anomalous:
        selected = [int(idx) for idx in selected if blended_scores[idx] > 0]

    return selected[:max_count]


def _correlation_from_covariance(covariance_matrix: np.ndarray) -> np.ndarray:
    cov = np.asarray(covariance_matrix, dtype=float)
    if cov.ndim != 2 or cov.shape[0] != cov.shape[1]:
        return cov

    variances = np.diag(cov)
    with np.errstate(divide="ignore", invalid="ignore"):
        inv_std = np.divide(
            1.0,
            np.sqrt(np.clip(np.abs(variances), 1e-12, None)),
            out=np.zeros_like(variances, dtype=float),
            where=np.abs(variances) > 1e-12,
        )

    corr = cov * np.outer(inv_std, inv_std)
    np.fill_diagonal(corr, 1.0)
    return corr


def build_output(
    df_processed: pd.DataFrame,
    id_cols: list,
    distances_sq: np.ndarray,
    estimates: np.ndarray,
    deviations: np.ndarray,
    outlier_flags: np.ndarray,
    observed_metrics: np.ndarray | None = None,
    mahalanobis_pvalue: np.ndarray | None = None,
    anomaly_score: np.ndarray | None = None,
    min_value_threshold: float = 0.0,
) -> pd.DataFrame:
    """
    Assemble the final output DataFrame, sorted by anomaly flag then Mahalanobis distance.

    Columns:
      <sanitized id columns> — cleaned ID fields
      <sanitized metric columns> — reported values
    mahalanobis_distance — actual Mahalanobis distance
    observed_metrics — number of non-missing metrics used for this row's MD
    mahalanobis_pvalue — right-tail p-value from chi2(observed_metrics)
    anomaly_score — -log10(mahalanobis_pvalue), larger means more anomalous
      anomalous — boolean flag
    top_deviation_indicator — metric column with the highest blended driver score
                      (normalized deviation + absolute residual, ignoring low values)
      E_<col> — estimated value for each metric column
    """
    original_metric_cols = [c for c in df_processed.columns if c not in id_cols]
    all_names = list(id_cols) + original_metric_cols
    sanitized_names = _sanitize_column_names(all_names)

    id_cols_clean = sanitized_names[: len(id_cols)]
    metric_cols_clean = sanitized_names[len(id_cols):]

    result = df_processed.copy().reset_index(drop=True)
    result.columns = sanitized_names
    result["_source_row"] = np.arange(len(result), dtype=int)

    result["mahalanobis_distance"] = np.sqrt(np.maximum(distances_sq, 0.0))
    if observed_metrics is not None:
        result["observed_metrics"] = np.asarray(observed_metrics, dtype=int)
    if mahalanobis_pvalue is not None:
        result["mahalanobis_pvalue"] = np.asarray(mahalanobis_pvalue, dtype=float)
    if anomaly_score is not None:
        result["anomaly_score"] = np.asarray(anomaly_score, dtype=float)
    result["anomalous"] = outlier_flags.astype(bool)

    data_vals = df_processed[original_metric_cols].values
    display_values = []
    for i in range(len(result)):
        row_values = []
        for j, col in enumerate(original_metric_cols):
            value = data_vals[i, j]
            est = estimates[i, j]
            if pd.isna(value):
                row_values.append("")
            elif np.isnan(est):
                row_values.append(str(value))
            else:
                row_values.append(f"{value:.6g} (exp {est:.6g})")
        display_values.append(row_values)

    display_df = pd.DataFrame(display_values, columns=metric_cols_clean)
    result[metric_cols_clean] = result[metric_cols_clean].astype(object)
    result.loc[:, metric_cols_clean] = display_df.astype(object).values

    result["top_deviation_indicator"] = None

    for i in range(len(result)):
        if not outlier_flags[i]:
            continue
        score_row, _ = _blended_driver_scores(
            row_devs=np.abs(np.asarray(deviations[i], dtype=float)),
            row_actuals=np.asarray(data_vals[i], dtype=float),
            row_expecteds=np.asarray(estimates[i], dtype=float),
            minimum_signal_scale=max(float(min_value_threshold), 1.0),
            minimum_absolute_delta=1.0,
            relative_delta_floor=0.10,
        )
        if np.nanmax(score_row) <= 0:
            continue
        result.at[i, "top_deviation_indicator"] = metric_cols_clean[int(np.argmax(score_row))]

    sort_cols = ["anomalous"]
    ascending = [False]
    if "anomaly_score" in result.columns:
        sort_cols.append("anomaly_score")
        ascending.append(False)
    sort_cols.append("mahalanobis_distance")
    ascending.append(False)

    result = result.sort_values(sort_cols, ascending=ascending).reset_index(drop=True)

    return result


def write_output(result: pd.DataFrame, report: dict, output_path: str) -> str:
    """Write the results as CSV or as an Excel workbook with explanatory tabs."""
    output_path = os.fspath(output_path)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    if not output_path.lower().endswith(".xlsx"):
        result.to_csv(output_path, index=False)
        return output_path

    metric_cols = report.get("metric_cols_used", [])
    analysis_data = report.get("analysis_data")
    estimated_data = report.get("estimated_data")
    covariance_matrix = report.get("covariance_matrix")
    deviations = report.get("deviations")
    id_cols = report.get("id_cols", [])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        metadata_cols = {
            "mahalanobis_distance",
            "observed_metrics",
            "mahalanobis_pvalue",
            "anomaly_score",
            "anomalous",
            "top_deviation_indicator",
            "_source_row",
        }
        sanitized_id_cols = set(_sanitize_column_names(id_cols))
        deviation_threshold = 0.25
        minimum_signal_scale = 1.0
        minimum_absolute_delta = 1.0
        relative_delta_floor = 0.10
        blended_threshold = 0.30
        min_highlights_anomalous = 1
        max_highlights_anomalous = 3

        export_cols = []
        for col in result.columns:
            if col in metadata_cols:
                export_cols.append(col)
            elif col in sanitized_id_cols:
                export_cols.append(col)
            elif col.startswith("e_") or col.startswith("d_") or col.startswith("D_"):
                continue
            else:
                export_cols.append(col)

        export_df = result.loc[:, export_cols].copy()
        export_df.to_excel(writer, sheet_name="facility_results", index=False)
        ws = writer.book["facility_results"]

        metric_col_positions = [
            idx for idx, col in enumerate(export_df.columns) if col not in metadata_cols and col not in sanitized_id_cols
        ]

        if deviations is not None:
            deviations_arr = np.asarray(deviations, dtype=float)
            analysis_array = np.asarray(analysis_data, dtype=float) if analysis_data is not None else None
            estimated_array = np.asarray(estimated_data, dtype=float) if estimated_data is not None else None
            for row_idx in range(2, len(export_df) + 2):
                row_num = row_idx - 2
                source_row = row_num
                if "_source_row" in result.columns and row_num < len(result):
                    try:
                        source_row = int(result.iloc[row_num]["_source_row"])
                    except (TypeError, ValueError):
                        source_row = row_num
                if source_row >= deviations_arr.shape[0] or source_row < 0:
                    continue
                row_devs = np.abs(np.asarray(deviations_arr[source_row], dtype=float))
                if row_devs.size == 0:
                    continue
                row_values = analysis_array[source_row] if analysis_array is not None and analysis_array.ndim == 2 and source_row < analysis_array.shape[0] else None
                row_expected_values = (
                    estimated_array[source_row]
                    if estimated_array is not None and estimated_array.ndim == 2 and source_row < estimated_array.shape[0]
                    else None
                )

                row_actuals = np.full(row_devs.shape[0], np.nan, dtype=float)
                row_expecteds = np.full(row_devs.shape[0], np.nan, dtype=float)
                for metric_idx, col_idx in enumerate(metric_col_positions):
                    if metric_idx >= row_devs.size:
                        continue
                    cell = ws.cell(row=row_idx, column=col_idx + 1)
                    if row_values is not None and metric_idx < row_values.size:
                        row_actuals[metric_idx] = _parse_numeric_value(row_values[metric_idx])
                    else:
                        row_actuals[metric_idx] = _parse_numeric_value(cell.value)

                    if row_expected_values is not None and metric_idx < row_expected_values.size:
                        row_expecteds[metric_idx] = _parse_numeric_value(row_expected_values[metric_idx])
                    else:
                        row_expecteds[metric_idx] = _parse_expected_from_display(cell.value)

                blended_scores, _ = _blended_driver_scores(
                    row_devs=row_devs,
                    row_actuals=row_actuals,
                    row_expecteds=row_expecteds,
                    minimum_signal_scale=minimum_signal_scale,
                    minimum_absolute_delta=minimum_absolute_delta,
                    relative_delta_floor=relative_delta_floor,
                )

                is_anomalous = False
                if "anomalous" in export_df.columns and row_num < len(export_df):
                    is_anomalous = bool(export_df.iloc[row_num]["anomalous"])

                finite_mask = np.isfinite(row_devs) & np.isfinite(row_actuals) & np.isfinite(row_expecteds)
                primary_mask = (
                    finite_mask
                    & (row_devs >= deviation_threshold)
                    & (blended_scores >= blended_threshold)
                )
                selected_indices = set(
                    _select_driver_indices(
                        blended_scores=blended_scores,
                        finite_mask=finite_mask,
                        primary_mask=primary_mask,
                        is_anomalous=is_anomalous,
                        min_count=min_highlights_anomalous,
                        max_count=max_highlights_anomalous,
                    )
                )

                for metric_idx, col_idx in enumerate(metric_col_positions):
                    if metric_idx >= row_devs.size:
                        continue
                    cell = ws.cell(row=row_idx, column=col_idx + 1)
                    deviation_value = row_devs[metric_idx]
                    if not np.isfinite(deviation_value):
                        cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
                        continue

                    blended_value = blended_scores[metric_idx] if metric_idx < blended_scores.size else 0.0
                    if metric_idx not in selected_indices or not np.isfinite(blended_value) or blended_value <= 0:
                        cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
                        continue

                    cell.fill = _fill_for_value(float(blended_value), 1.0)

        if covariance_matrix is not None and metric_cols:
            corr_matrix = _correlation_from_covariance(covariance_matrix)
            corr_df = pd.DataFrame(corr_matrix, index=metric_cols, columns=metric_cols)
            corr_df.to_excel(writer, sheet_name="correlation_matrix")
            ws = writer.book["correlation_matrix"]
            max_abs = max(abs(float(v)) for v in corr_df.to_numpy().ravel() if pd.notna(v) and np.isfinite(v)) if corr_df.size else 0.0
            for row_idx in range(2, len(metric_cols) + 2):
                for col_idx in range(2, len(metric_cols) + 2):
                    value = ws.cell(row=row_idx, column=col_idx).value
                    if isinstance(value, (int, float)) and np.isfinite(value):
                        ws.cell(row=row_idx, column=col_idx).fill = _fill_for_value(float(value), max_abs)

        if analysis_data is not None and metric_cols:
            analysis_array = np.asarray(analysis_data)
            ranges_df = pd.DataFrame(
                {
                    "variable": metric_cols,
                    "min": np.nanmin(analysis_array, axis=0),
                    "max": np.nanmax(analysis_array, axis=0),
                    "range": np.nanmax(analysis_array, axis=0) - np.nanmin(analysis_array, axis=0),
                    "mean": np.nanmean(analysis_array, axis=0),
                    "std": np.nanstd(analysis_array, axis=0),
                }
            )
            ranges_df.to_excel(writer, sheet_name="variable_ranges", index=False)

    return output_path
